import json
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader


MIME_BY_EXTENSION = {
    ".pdf": "application/pdf",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".webp": "image/webp",
    ".txt": "text/plain",
    ".html": "text/html",
    ".json": "application/json",
    ".zip": "application/zip",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


class OutputValidationError(ValueError):
    """Raised when a converter produces an unusable output file."""


def validate_output(path: Path, *, expected_extension: str, expected_mime: str, max_bytes: int | None = None) -> dict:
    if not path.exists() or not path.is_file() or path.stat().st_size == 0:
        raise OutputValidationError("لم يُنتج محرك التحويل ملفًا صالحًا.")

    if max_bytes is not None and path.stat().st_size > max_bytes:
        raise OutputValidationError("حجم الملف الناتج يتجاوز الحد الآمن.")

    extension = path.suffix.lower()
    if extension != expected_extension.lower():
        raise OutputValidationError("امتداد الملف الناتج لا يطابق العملية المطلوبة.")

    expected_by_extension = MIME_BY_EXTENSION.get(extension)
    if expected_by_extension and expected_by_extension != expected_mime:
        raise OutputValidationError("نوع الملف الناتج غير متوافق مع امتداده.")

    if extension == ".pdf":
        try:
            pages = len(PdfReader(str(path), strict=False).pages)
        except Exception as exc:
            raise OutputValidationError("ملف PDF الناتج غير صالح.") from exc
        if pages < 1:
            raise OutputValidationError("ملف PDF الناتج لا يحتوي على صفحات.")
        return {"pages": pages}

    if extension in {".jpg", ".jpeg", ".png", ".webp"}:
        try:
            with Image.open(path) as image:
                image.verify()
                width, height = image.size
        except Exception as exc:
            raise OutputValidationError("ملف الصورة الناتج غير صالح.") from exc
        if width < 1 or height < 1:
            raise OutputValidationError("أبعاد الصورة الناتجة غير صالحة.")
        return {"width": width, "height": height}

    if extension == ".txt":
        try:
            text = path.read_text(encoding="utf-8")
        except Exception as exc:
            raise OutputValidationError("ملف النص الناتج غير صالح.") from exc
        if not text.strip():
            raise OutputValidationError("ملف النص الناتج فارغ.")
        return {"characters": len(text)}

    if extension == ".html":
        try:
            text = path.read_text(encoding="utf-8")
        except Exception as exc:
            raise OutputValidationError("ملف HTML الناتج غير صالح.") from exc
        if "<html" not in text.lower():
            raise OutputValidationError("ملف HTML الناتج غير صالح.")
        return {"characters": len(text)}

    if extension == ".json":
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception as exc:
            raise OutputValidationError("ملف JSON الناتج غير صالح.") from exc
        return {"keys": len(data) if isinstance(data, dict) else 0}

    if extension == ".zip":
        try:
            with zipfile.ZipFile(path) as zf:
                bad = zf.testzip()
                names = zf.namelist()
        except Exception as exc:
            raise OutputValidationError("ملف ZIP الناتج غير صالح.") from exc
        if bad is not None or not names:
            raise OutputValidationError("ملف ZIP الناتج تالف أو فارغ.")
        return {"entries": len(names)}

    if extension == ".xlsx":
        try:
            workbook = load_workbook(path, read_only=True)
            sheet_count = len(workbook.sheetnames)
            workbook.close()
        except Exception as exc:
            raise OutputValidationError("ملف XLSX الناتج غير صالح.") from exc
        if sheet_count < 1:
            raise OutputValidationError("ملف XLSX الناتج لا يحتوي على أوراق عمل.")
        return {"sheets": sheet_count}

    raise OutputValidationError("نوع الملف الناتج غير مدعوم للتحقق.")
